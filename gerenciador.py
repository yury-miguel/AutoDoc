# *-* coding: utf-8 *-*

import os
import time
import shutil
import difflib
import pdf_auto
import openpyxl
import email_auto
from datetime import datetime
from dotenv import load_dotenv, dotenv_values

# CARREGA AS VARIAVEIS DE AMBIENTE E O ARMAZENAMENTO GLOBAL DOS ARQUIVOS
load_dotenv()


# FUNÇÃO QUE ATUALIZA O VALOR DAS CONFIGURAÇÕES DO SISTEMA
def atualiza_env(documentos, downloads, email_padrao, senha_padrao, assunto_padrao, mensagem_padrao, ano_docs):
    try:
        env_vars = {}

        with open('.env', 'r', encoding='utf-8') as arquivo:
            for linha in arquivo:
                if linha.strip():
                    chave, valor = linha.strip().split('=', 1)
                    env_vars[chave] = valor.strip()
        
            env_vars["DOWN"] = downloads.strip()
            env_vars["DOCS"] = documentos.strip()
            env_vars["EMAIL"] = email_padrao.strip()
            env_vars["SENHA"] = senha_padrao.strip()
            env_vars["ASSUNTO"] = assunto_padrao.strip()
            env_vars["MENSAGEM"] = mensagem_padrao.strip()
            env_vars["ANO"] = ano_docs.strip()
        
        with open('.env', 'w', encoding='utf-8') as arquivo:
            for chave, valor in env_vars.items():
                arquivo.write(f"{chave}={valor}\n")

        load_dotenv()

    except Exception as e:
        registros(f"Erro ao atualizar as variáveis de ambiente: {e}")



# FUNÇÃO QUE CADASTRA OS CLIENTES E SEUS EMAILS
def cadastra_dados(nome, email):
    try:
        arquivo_excel = 'clientes.xlsx'

        if not os.path.exists(arquivo_excel):
            livro = openpyxl.Workbook()
            folha = livro.active
            folha.title = "Clientes"
            folha.append(["Nome", "Email"])
            livro.save(arquivo_excel)
        
        livro = openpyxl.load_workbook(arquivo_excel)
        folha = livro.active
        folha.append([nome, email])
        livro.save(arquivo_excel)

    except Exception as e:
        registros(f"Erro ao cadastrar dados no arquivo {arquivo_excel}: {e}")



# FUNÇÃO QUE CONTROLA O ENVIO DE EMAILS MANUAIS
def emails_manuais(destinatario, assunto, mensagem, caminho_anexo=None):
    load_dotenv()
    dados = dotenv_values(".env")
    remetente = dados['EMAIL']
    senha = dados['SENHA']

    try:
        if remetente and senha is not None:
            email = email_auto.enviar_email(remetente, senha, destinatario, assunto, mensagem, caminho_anexo)
            return email
        else:
            raise ValueError('Remetente ou senha não definidos.')
    
    except Exception as e:
        registros(f"Erro ao enviar email manual para {destinatario}: {e}")
        return 'Erro ao Enviar Email'



# FUNÇÃO QUE CONTROLA O ENVIO DE EMAILS AUTOMÁTICOS
def emails_automaticos(destinatario_nome=None, caminho_anexo=None):
    load_dotenv()
    dados = dotenv_values(".env")
    remetente = dados['EMAIL']
    senha = dados['SENHA']
    assunto = dados['ASSUNTO']
    mensagem = dados['MENSAGEM']

    try:
        if remetente and senha and destinatario_nome is not None:
            destinatario_email = buscar_email_por_nome(destinatario_nome)

            if destinatario_email and caminho_anexo is not None:
                anexo_nome = os.path.basename(caminho_anexo)

                with open(caminho_anexo, 'rb') as anexo:
                    email_auto.enviar_email(remetente, senha, destinatario_email, assunto, mensagem, anexo, anexo_nome)
            else:
                raise ValueError(f"Email ou anexo não encontrado para o destinatário: {destinatario_nome}:{caminho_anexo}")
        else:
            raise ValueError('Remetente, senha ou destinatário não definidos.')
    
    except Exception as e:
        registros(f"Erro ao enviar email automático para {destinatario_nome}: {e}")



# FUNÇÃO QUE BUSCA EMAIL POR NOME NA PLANILHA DE CLIENTES
def buscar_email_por_nome(nome):
    try:
        arquivo_excel = 'clientes.xlsx'

        if not os.path.exists(arquivo_excel):
            raise FileNotFoundError("Arquivo 'clientes.xlsx' não encontrado.")

        livro = openpyxl.load_workbook(arquivo_excel)
        folha = livro.active

        for linha in folha.iter_rows(min_row=2, values_only=True):
            nome_excel, email_excel = linha
            if nome_excel == nome:
                return email_excel

        return None

    except Exception as e:
        registros(f"Erro ao buscar email por nome '{nome}': {e}")
        return None



# REGISTRA O STATUS DOS PDFS
def status_pdf(nome_pdf, categoria, nome_extraido, caminho):
    try:
        with open('status_pdfs.txt', 'a', encoding='utf-8') as log_file:
            data_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"Arquivo PDF: {nome_pdf}\n")
            log_file.write(f"Categoria: {categoria}\n")
            log_file.write(f"Nome Extraído: {nome_extraido}\n")
            log_file.write(f"Caminho: {caminho}\n")
            log_file.write(f"Data e Hora: {data_hora}\n")
            log_file.write(f"{'-'*40}\n\n")

    except Exception as e:
        registros(f"Erro ao registrar status do PDF {nome_pdf}: {e}")



# MOVE OS PDFS BAIXADOS PARA SUAS DEVIDAS PASTAS
def mover_pdf_para_pasta(nome_pdf, categoria, nome_extraido, caminho_pdf):
    try:
        load_dotenv()
        dados = dotenv_values(".env")
        caminho_docs = dados["DOCS"]
        ano_docs = dados["ANO"]

        subpastas = ["MEI", "LUCRO PRESUMIDO", "SIMPLES NACIONAL"]
        empresa_encontrada = False
        similaridade_limite = 0.5

        for subpasta in subpastas:
            caminho_subpasta = os.path.join(caminho_docs, subpasta)
            empresas_existentes = os.listdir(caminho_subpasta)
            nome_proximo = difflib.get_close_matches(nome_extraido, empresas_existentes, n=1, cutoff=similaridade_limite)
            
            if nome_proximo:
                caminho_empresa = os.path.join(caminho_subpasta, nome_proximo[0])

                if os.path.exists(caminho_empresa):
                    empresa_encontrada = True

                    caminho_categoria = os.path.join(caminho_empresa, categoria)
                    if not os.path.exists(caminho_categoria):
                        os.makedirs(caminho_categoria)

                    caminho_ano = os.path.join(caminho_categoria, ano_docs)
                    if not os.path.exists(caminho_ano):
                        os.makedirs(caminho_ano)

                    novo_caminho_pdf = os.path.join(caminho_ano, nome_pdf)
                    shutil.move(caminho_pdf, novo_caminho_pdf)
                    return novo_caminho_pdf

        if not empresa_encontrada:
            caminho_empresa_sistema = os.path.join(caminho_docs, f"Sistema_{nome_extraido}")

            if not os.path.exists(caminho_empresa_sistema):
                os.makedirs(caminho_empresa_sistema)

            caminho_categoria = os.path.join(caminho_empresa_sistema, categoria)
            if not os.path.exists(caminho_categoria):
                os.makedirs(caminho_categoria)

            caminho_ano = os.path.join(caminho_categoria, ano_docs)
            if not os.path.exists(caminho_ano):
                os.makedirs(caminho_ano)

            novo_caminho_pdf = os.path.join(caminho_ano, nome_pdf)
            shutil.move(caminho_pdf, novo_caminho_pdf)
            return novo_caminho_pdf

    except Exception as e:
        registros(f"Erro ao mover PDF {nome_pdf} para pasta {categoria}: {e}")



# FUNÇÃO QUE CONTROLA E GERENCIA OS PDFS BAIXADOS
def monitorar_pdfs():
    load_dotenv()
    dados = dotenv_values(".env")
    caminho_down = dados["DOWN"]
    registrado = False

    while True:
        try:
            if not caminho_down:
                if registrado == False:
                    registros("Erro: O caminho de downloads não está configurado.")
                    registrado = True
                    
                load_dotenv()
                dados = dotenv_values(".env")
                caminho_down = dados["DOWN"]
                time.sleep(5)
                continue

            arquivos_pdfs = [f for f in os.listdir(caminho_down) if f.endswith('.pdf')]

            for arquivo in arquivos_pdfs:
                caminho_pdf = os.path.join(caminho_down, arquivo)
                extrator = pdf_auto.ExtracaoPdf(caminho_pdf)
                
                if extrator.nome and extrator.categoria:
                    novo_caminho = mover_pdf_para_pasta(arquivo, extrator.categoria, extrator.nome, caminho_pdf)
                    emails_automaticos(destinatario_nome=extrator.nome, caminho_anexo=novo_caminho)
                    status_pdf(arquivo, extrator.categoria, extrator.nome, novo_caminho)
                else:
                    status_pdf(arquivo, extrator.categoria, extrator.nome, caminho_pdf)

            time.sleep(3)
        
        except Exception as e:
            registros(f"Erro ao monitorar arquivos: {e}")
            time.sleep(5)



# RETORNA OS 6 ULTIMOS STATUS DOS PDFS PROCESSADOS E BAIXADOS
def ultimos_dados():
    try:    
        if not os.path.exists('status_pdfs.txt'):
            return []
        
        with open('status_pdfs.txt', 'r', encoding='utf-8') as arquivo:
            linhas = arquivo.readlines()
        
        entradas = []
        dados_atual = {}

        for linha in reversed(linhas):
            linha = linha.strip()
            
            if linha.startswith('Arquivo PDF:'):
                if dados_atual:
                    if 'caminho' in dados_atual and ('categoria' in dados_atual or 'nome_extraido' in dados_atual):
                        entradas.append(dados_atual)
                    dados_atual = {}
                
                dados_atual['status'] = 'Não Verificado'
            elif linha.startswith('Categoria:'):
                dados_atual['categoria'] = linha.replace('Categoria: ', '')
            elif linha.startswith('Nome Extraído:'):
                dados_atual['nome_extraido'] = linha.replace('Nome Extraído: ', '')
            elif linha.startswith('Caminho:'):
                dados_atual['caminho'] = linha.replace('Caminho: ', '')
            elif linha.startswith('Data e Hora:'):
                dados_atual['data_hora'] = linha.replace('Data e Hora: ', '')
        
        if dados_atual:
            if 'caminho' in dados_atual and ('categoria' in dados_atual or 'nome_extraido' in dados_atual):
                entradas.append(dados_atual)
        
        ultimos_6 = entradas[:6]
        
        for entrada in ultimos_6:
            if entrada.get('categoria') and entrada.get('nome_extraido') and entrada['categoria'] != 'None' and entrada['nome_extraido'] != 'None':
                entrada['status'] = 'Verificado'
            else:
                entrada['status'] = 'Não Verificado'
          
        return ultimos_6
    
    except Exception as e:
        registros(f"Erro ao ler o arquivo status_pdfs.txt: {e}")
        return []



# FUNÇÃO QUE FILTRA DOCUMENTOS
def filtrar_dados(filtro):
    try:
        if not os.path.exists('status_pdfs.txt'):
            return []

        with open('status_pdfs.txt', 'r', encoding='utf-8') as arquivo:
            linhas = arquivo.readlines()

        entradas = []
        dados_atual = {}

        for linha in linhas:
            linha = linha.strip()

            if linha.startswith('Arquivo PDF:'):
                if dados_atual:
                    if dados_atual.get('categoria') and dados_atual.get('nome_extraido'):
                        dados_atual['status'] = 'verificado'
                    else:
                        dados_atual['status'] = 'não verificado'
                    entradas.append(dados_atual)
                    dados_atual = {}

                dados_atual['caminho'] = linha.replace('Arquivo PDF: ', '')

            elif linha.startswith('Categoria:'):
                dados_atual['categoria'] = linha.replace('Categoria: ', '')

            elif linha.startswith('Nome Extraído:'):
                dados_atual['nome_extraido'] = linha.replace('Nome Extraído: ', '')

        if dados_atual:
            if dados_atual.get('categoria') and dados_atual.get('nome_extraido'):
                dados_atual['status'] = 'verificado'
            else:
                dados_atual['status'] = 'não verificado'
            entradas.append(dados_atual)

        filtro = filtro.lower()
        filtrados = [
            entrada for entrada in entradas
            if filtro in entrada.get('categoria', '').lower() or
               filtro in entrada.get('nome_extraido', '').lower() or
               filtro in entrada.get('caminho', '').lower()
        ]

        return filtrados

    except Exception as e:
        registros(f"Erro ao ler e filtrar o arquivo status_pdfs.txt: {e}")
        return []



# FUNÇÃO QUE REGISTRA AS OCORRENCIAS DURANTE FUNCIONAMENTO DO SISTEMA
def registros(texto):
    data_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open('logs.txt', 'a', encoding='utf-8') as arquivo_log:
        arquivo_log.write(f"{data_hora} - {texto}\n")



# FUNÇÃO QUE RETORNA OS REGISTROS DE FORMA ORGANIZADA PARA SEREM EXIBIDOS
def retorna_registros(data_filtro=None):
    try:
        if not os.path.exists('logs.txt'):
            return []

        with open('logs.txt', 'r', encoding='utf-8') as arquivo:
            linhas = arquivo.readlines()

        if data_filtro:
            registros_filtrados = [
                linha for linha in linhas if linha.startswith(data_filtro)
            ]
        else:
            registros_filtrados = linhas

        return registros_filtrados

    except Exception as e:
        registros(f"Erro ao ler logs: {e}")
        return []